import openpyxl as op
import re

ExcelFile = "./Split Address.xlsx"

#This is the column index of Data in Excel
Col_Add1 = 3
Col_Add2 = 4
Col_Building = 6
Col_Unit = 7
Col_StreetNo = 8
Col_StreetName = 9
Col_StreetType = 10

#This is the column index of Street Type Mapping in Excel
Col_AllType = 1
Col_StrType = 3
Col_StrAbbr = 4


def SplitAddress(wsData, wsType):
    wsMaxRow = wsData.max_row
    TypeMaxRow = len(wsType['A'])
    AllStrType=[]

    #Populate all Street Type for Regex later
    for i in range(2,TypeMaxRow):
        AllStrType.append(wsType.cell(row=i, column=Col_AllType).value.strip())

    #Run through every Excel Row to the last row
    for currow in range(2, wsMaxRow+1):
        Add1 = wsData.cell(currow, Col_Add1).value if wsData.cell(currow, Col_Add1).value != None else ""
        Add2 = wsData.cell(currow, Col_Add2).value if wsData.cell(currow, Col_Add2).value != None else ""
        #Combine two addresses from Excel together
        FullAdress = (Add1 + " " + Add2).strip()

        #-------------------------------------------------------------------------------------
        #Regex for Unit, and return the Unit out from the string
        UnitRegex=[]
        UnitRegex.append("(U)(nit)?(\s+)?[0-9]+(\s[A-Da-d]\s)?")
        UnitRegex.append("(N)(umber|o)?(\s+)?[0-9]+(\s[A-Da-d]\s)?")
        UnitRegex.append("[0-9]+[A-Da-d]?(\s+)?[/]")
        for item in UnitRegex:
            Unit = MatchedRegex(FullAdress, item)
            if Unit != "": break

        #Remove the matched Unit out of the Full Address
        FullAdress = RemoveStr(FullAdress, Unit).strip()

        #Remove the unnecessary / in the Unit
        Unit = RemoveStr(Unit, r"/").strip()

        #-------------------------------------------------------------------------------------
        #Regex for Full Street
        StreetRegex = r"([0-9]+[\s]?[-]\s?)?[0-9]+[abAB]?[\s][a-zA-Z\s]+(\s)?"
        FullStreet = FullAdress
        StreetType =""
        for item in AllStrType:
            NewStreetRegex = StreetRegex + item
            FullStreet = MatchedRegex(FullAdress, NewStreetRegex)
            if FullStreet != "":
                StreetType = item
                break

        #Regex for Street Number
        StreetNoRegex = r"([0-9]+[\s]?[-]\s?)?[0-9]+[abAB]?"
        StreetNo = MatchedRegex(FullStreet,StreetNoRegex).strip()

        #Regex for Street Name
        StreetNameRegex = r"[a-zA-Z\s]+(\s)?(?=" + StreetType + ")"
        StreetName = MatchedRegex(FullStreet, StreetNameRegex).strip()

        #-------------------------------------------------------------------------------------
        #Remove Full Street from Address is Building Name
        Building = RemoveStr(FullAdress, FullStreet).strip()
        #Remove unnecessary comma in Building
        CommaCount = Building.count(",")
        while CommaCount!=0:
            Building=RemoveStr(Building,",")
            CommaCount -= 1

        # -------------------------------------------------------------------------------------
        #Populate data into Excel
        wsData.cell(currow, Col_Building).value = Building
        wsData.cell(currow, Col_Unit).value = Unit
        wsData.cell(currow, Col_StreetNo).value = StreetNo
        wsData.cell(currow, Col_StreetName).value = StreetName
        wsData.cell(currow, Col_StreetType).value = StreetType

    wb.save(ExcelFile)



def RemoveStr(FullStr, RemovedStr):
    IsFound = FullStr.find(RemovedStr)
    ReturnStr = FullStr

    #Get the string after the length of removed string if the match is in first position
    if IsFound == 0:
        RightStr = FullStr[len(RemovedStr):].strip()
        ReturnStr = RightStr

    #If not at the first position, need to break down to left and right, and then concatenate them
    if IsFound > 0:
        RightPosition = len(RemovedStr) + IsFound
        LeftStr = FullStr[:IsFound].strip()
        RightStr = FullStr[RightPosition:].strip()
        ReturnStr = LeftStr + " " + RightStr

    return ReturnStr

def MatchedRegex(InputStr, Regex):
    Match = re.search(Regex, InputStr)
    return (Match.group() if Match != None else "")

if __name__ == '__main__':
    wb = op.load_workbook(ExcelFile, data_only=True)
    wsData = wb["Data"]
    wsType = wb["StreetType"]
    SplitAddress(wsData, wsType)





