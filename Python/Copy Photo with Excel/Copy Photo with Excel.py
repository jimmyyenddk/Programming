import os
import openpyxl as op
import shutil

#This is to map the location of Excel file, location of the photo, and the location that need to be pasted to
StaffExcelList = r'Staff List.xlsx'
PhotoCopyPath = r'.\Photo'
DestinationPath =r'.\Destination'

#This is the column index in Excel worksheet
StaffIDCol = 1
CopiedResultCol = 5
ResultPathCol = 6
Morethan1Col = 7

def CopyPhoto():
    wb = op.load_workbook(StaffExcelList, data_only=True)
    ws = wb["Data"]
    ExcelMaxRow = ws.max_row

    #This is a list of image file extension, can be add more for all image type extension
    FileExtList = [".png", ".jpg", ".jpeg", ".bmp"]

    #This is to run through all folders and subfolders in provided path
    for root, subdirs, files in os.walk(PhotoCopyPath):
        for file in files:
            if file.endswith(tuple(FileExtList)):
                #If file extension in the list provided, then it will split the file name and extension type
                FileName, FileExt = os.path.splitext(file)

                for ExcelRow in range(2,ExcelMaxRow):
                    StaffID = ws.cell(row = ExcelRow, column = StaffIDCol).value

                    #Run through all the row in Excel and check whether the StaffID is in the filename
                    if StaffID in FileName:
                        #If two files have the same name, it will copy the latest
                        #This column is to flag whether it has been overriden.
                        if ws.cell(row=ExcelRow, column=CopiedResultCol).value == "Yes":
                            ws.cell(row=ExcelRow, column= Morethan1Col).value = "Yes"

                        #This is to update the Excel File
                        ws.cell(row=ExcelRow, column=CopiedResultCol).value = "Yes"
                        ws.cell(row=ExcelRow, column=ResultPathCol).value = root

                        #Create a path to be ready for copy and paste photos
                        CurrentPath = os.path.join(root,file)
                        NewFileName = StaffID + FileExt
                        PastedPath = os.path.join(DestinationPath, NewFileName)
                        shutil.copyfile(CurrentPath,PastedPath)

                        #Once copy is done, it will break the loop for in the Excel list, and move to different file.
                        break

    #Save the Excel File
    wb.save(StaffExcelList)

if __name__ == '__main__':
    CopyPhoto()
