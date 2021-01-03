import os
import openpyxl as op

SummaryExcel = "./Summary.xlsx"
FileLocation = "./Excel Files"
#ClientPrefix = ["PRO", "BUR", "MAN"]


def leftstring(str, no):
    return str[:no]


def getsummary(ws, fileloc):
    SummaryMaxrow = ws.max_row
    ExcelCurRow = 1

    for root, subdirs, files in os.walk(fileloc):
        for file in files:
            if file.endswith(".xlsx"):
                FilePath = os.path.join(root, file)  # To get full folder path of this Excel File
                ExcelFile = op.load_workbook(FilePath, data_only= True)  # Then load the file and assign variable

                for sheet in ExcelFile.worksheets:
                    LastRow = sheet.max_row

                    RentInc = ProInc = GovInc = IntInc = 0
                    ProExp = LicExp = BankExp = OtherExp = 0
                    TTInc = TTExp = OB = CB = 0
                    ClientID = Month = None

                    TTIncRow = searchstring(sheet, "total income")
                    TTExpRow = searchstring(sheet, "total expense")

                    for row in range(2, LastRow):
                        if TTIncRow == 0 or TTExpRow == 0:
                            break  # Move to other sheet as current sheet is not a statement

                        TextCol = sheet.cell(row, 1).value.lower() if sheet.cell(row, 1).value != None else "" # Text in column 1
                        ValueCol = sheet.cell(row, 6).value  # All value in column 6

                        TTInc = sheet.cell(TTIncRow, 6).value  # Get a total income
                        TTExp = sheet.cell(TTExpRow, 6).value  # Get a total expense

                        if row < TTIncRow:
                            if TextCol == "client id":
                                ClientID = sheet.cell(row, 2).value  # Get ClientID

                            datevalue = sheet.cell(row, 5).value
                            if datevalue != None and datevalue.lower() == "date":
                                Month = ValueCol  # Get the date in the current tab

                            if "rent inc" in TextCol:
                                RentInc += ValueCol  # To add all rent income together
                            if "profession" in TextCol:
                                ProInc += ValueCol  # To add all work income together
                            if "government" in TextCol:
                                GovInc += ValueCol  # To add all government sub income  together
                            if "interest" in TextCol:
                                IntInc += ValueCol  # To add all interest income together
                            if "opening bal" in TextCol:
                                OB = ValueCol  # Get a Opening Balance
                        else:
                            if "property" in TextCol:
                                ProExp += ValueCol  # To add all property expense together
                            if "license" in TextCol:
                                LicExp += ValueCol  # To add all license expense together
                            if "banking" in TextCol:
                                BankExp += ValueCol  # To add all banking expense together
                            if "other" in TextCol:
                                OtherExp += ValueCol  # To add all other expense together
                            if "closing bal" in TextCol:
                                CB = ValueCol  # Get a closing balance value

                    # ===================================================================
                    # Populate these figures to Excel File
                    SummaryMaxrow += 1

                    ws.cell(SummaryMaxrow, 1).value = ClientID
                    ws.cell(SummaryMaxrow, 2).value = Month
                    ws.cell(SummaryMaxrow, 2).number_format = "mm/yyyy"
                    ws.cell(SummaryMaxrow, 3).value = RentInc
                    ws.cell(SummaryMaxrow, 4).value = ProInc
                    ws.cell(SummaryMaxrow, 5).value = GovInc
                    ws.cell(SummaryMaxrow, 6).value = IntInc
                    ws.cell(SummaryMaxrow, 7).value = ProExp
                    ws.cell(SummaryMaxrow, 8).value = LicExp
                    ws.cell(SummaryMaxrow, 9).value = BankExp
                    ws.cell(SummaryMaxrow, 10).value = OtherExp
                    ws.cell(SummaryMaxrow, 11).value = TTInc
                    ws.cell(SummaryMaxrow, 12).value = TTExp
                    ws.cell(SummaryMaxrow, 13).value = OB
                    ws.cell(SummaryMaxrow, 14).value = CB


def searchstring(sheet, str):
    result = 0
    max_row = sheet.max_row

    for row in range(2, max_row):
        SearchedText = sheet.cell(row, 1).value
        if SearchedText !=None and str in SearchedText.lower():
            result = row
            break
    return result

if __name__ == '__main__':
    wb = op.load_workbook(SummaryExcel, data_only= True)
    ws = wb["Summary"]
    getsummary(ws, FileLocation)
    wb.save(SummaryExcel)
